# from pandas import read_excel
from openpyxl import load_workbook
import os
import math
import json
from itertools import combinations

if os.getcwd().endswith('src'):
    data_root = '../data/'
    output_folder = '../outputs/'
    report_folder = '../report_rc/'
    model_folder = '../models/'
else:
    data_root = './data/'
    output_folder = './outputs/'
    report_folder = './report_rc/'
    model_folder = './models/'

station_list = ['aotizhongxin_aq', 'badaling_aq', 'beibuxinqu_aq', 'daxing_aq',
                'dingling_aq', 'donggaocun_aq', 'dongsi_aq', 'dongsihuan_aq',
                'fangshan_aq', 'fengtaihuayuan_aq', 'guanyuan_aq', 'gucheng_aq',
                'huairou_aq', 'liulihe_aq', 'mentougou_aq', 'miyun_aq',
                'miyunshuiku_aq', 'nansanhuan_aq', 'nongzhanguan_aq',
                'pingchang_aq', 'pinggu_aq', 'qianmen_aq', 'shunyi_aq',
                'tiantan_aq', 'tongzhou_aq', 'wanliu_aq', 'wanshouxigong_aq',
                'xizhimenbei_aq', 'yanqin_aq', 'yizhuang_aq', 'yongdingmennei_aq',
                'yongledian_aq', 'yufa_aq', 'yungang_aq', 'zhiwuyuan_aq']

# def read_locale(file_name):
    
#     def extract_info(row_id, booksheet):
#         station_name = booksheet.cell(row=row_id, column=1).value
#         longtitude = booksheet.cell(row=row_id, column=2).value
#         latitude = booksheet.cell(row=row_id, column=3).value
#         return station_name, longtitude, latitude

#     file_path = os.path.join(data_root, file_name)
#     xls_workbook = load_workbook(file_path)

#     work_sheet = xls_workbook.get_sheet_by_name(xls_workbook.get_sheet_names()[0])
#     # xls_rows = work_sheet.rows
#     locale_info = {}

#     for i in range(13, 25):     # Urban station: 0
#         info = extract_info(i, work_sheet)
#         locale_info[info[0]] = [(info[1], info[2]), 'Urban']

#     for i in range(27, 38):
#         info = extract_info(i, work_sheet)
#         locale_info[info[0]] = [(info[1], info[2]), 'SubUrban']

#     for i in range(40, 47):
#         info = extract_info(i, work_sheet)
#         locale_info[info[0]] = [(info[1], info[2]), 'Others']

#     for i in range(49, 54):
#         info = extract_info(i, work_sheet)
#         locale_info[info[0]] = [(info[1], info[2]), 'NearTraffic']
        
#     return locale_info

# def gen_distance_map(station_locale):
#     '''
#     return dist_map = {'station1-station2': distance, ...}
#     '''

#     def cal_dis(longitude1, latitude1, longitude2, latitude2):
#         latitude1 = (math.pi/180)*latitude1
#         latitude2 = (math.pi/180)*latitude2
#         longitude1 = (math.pi/180)*longitude1
#         longitude2= (math.pi/180)*longitude2
#         R = 6378.1
#         d =  math.acos(math.sin(latitude1)*math.sin(latitude2)+\
#             math.cos(latitude1)*math.cos(latitude2)*math.cos(longitude2-longitude1))*R
#         return d  # unit: km

#     dist_map = {}

#     stations = list(station_locale.keys())
#     for station_pair in combinations(stations, 2):
#         longitude1, latitude1 = station_locale[station_pair[0]][0]
#         longitude2, latitude2 = station_locale[station_pair[1]][0]
#         dist = cal_dis(longitude1, latitude1, longitude2, latitude2)
#         stations = sorted(list(station_pair))
#         dist_map[f'{stations[0]}-{stations[1]}'] = dist
#     return dist_map

# def gen_angle_map(station_locale):
#     '''
#     return dist_map = {'station1-station2': distance, ...}
#     '''

#     def cal_angle(longitude1, latitude1, longitude2, latitude2):
#         x = longitude2-longitude1
#         y = latitude2-latitude1
#         if y == 0:
#             angle = 90.0
#         else:
#             tan = x / y
#             angle = math.atan(tan) * 180 / math.pi

#         # The first quadrant
#         if y >= 0 and x >= 0:
#             pass
#         elif y < 0 and x >= 0:
#             angle += 180
#         elif y <= 0 and x < 0:
#             angle += 180
#         elif y > 0 and x < 0:
#             angle += 270  # unit: km
#         return angle

#     angle_map = {}

#     stations = list(station_locale.keys())
#     for station_pair in combinations(stations, 2):
#         longitude1, latitude1 = station_locale[station_pair[0]][0]
#         longitude2, latitude2 = station_locale[station_pair[1]][0]
#         angle = cal_angle(longitude1, latitude1, longitude2, latitude2)
#         stations = station_pair
#         angle_map[f'{stations[0]}-{stations[1]}'] = angle
#         angle_map[f'{stations[1]}-{stations[0]}'] = (angle + 180.0) % 360
#     return angle_map

# station_locale = read_locale('Beijing_AirQuality_Stations_en.xlsx')
# station_dist = gen_distance_map(station_locale)
# station_included_angles = gen_angle_map(station_locale)

# with open(os.path.join(output_folder, 'station_locales.json'), 'w') as f:
#     json.dump(station_locale, f)
# with open(os.path.join(output_folder, 'station_distances.json'), 'w') as f:
#     json.dump(station_dist, f)
# with open(os.path.join(output_folder, 'station_included_angles.json'), 'w') as f:
#     json.dump(station_included_angles, f)

with open(os.path.join(output_folder, 'station_locales.json'), 'r') as f:
    station_locale = json.load(f)
with open(os.path.join(output_folder, 'station_distances.json'), 'r') as f:
    station_dist = json.load(f)
with open(os.path.join(output_folder, 'station_included_angles.json'), 'r') as f:
    station_angle = json.load(f)



# print('done')